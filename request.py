import os
import re
import openpyxl
def trans(row):
    return "https://www.qdxq.sdu.edu.cn/" + row[0]
#python使人头大
import requests
import pandas as pd

url='https://www.qdxq.sdu.edu.cn/xwzx1/xwjj/xqyw.htm'
resp=requests.get(url=url)
temp=resp.content.decode()
obj=re.compile(r"<li>.*?/../(?P<aaa>info/\w{4}/\w{5}.htm).*?<span>(?P<bbb>.*?)</span>.*?<span>(?P<ccc>.*?)</span>.*?\"title\">(?P<ddd>.*?)</p>.*?\"content\">(?P<eee>.*?)</p>.*?</div>",re.S)
pri=obj.findall(temp)
print(pri)
data={
"context":["website","date","month","title","main idea"],
"idea0":pri[0],
"idea1":pri[1],
"idea2":pri[2],
"idea3":pri[3],
"idea4":pri[4],
"idea5":pri[5],
"idea6":pri[6],
"idea7":pri[7],
"idea8":pri[8],
"idea9":pri[9]
    }
df=pd.DataFrame(data)
df1=df.T
for i in df:
    print(i)
df1.to_excel("校区要闻.xlsx",index=False)
wb=openpyxl.load_workbook('校区要闻.xlsx')
sheet=wb['Sheet1']
sheet['A12']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A12'].value)
sheet['A3']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A3'].value)
sheet['A4']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A4'].value)
sheet['A5']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A5'].value)
sheet['A6']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A6'].value)
sheet['A7']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A7'].value)
sheet['A8']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A8'].value)
sheet['A9']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A9'].value)
sheet['A10']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A10'].value)
sheet['A11']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A11'].value)
wb.save('校区要闻.xlsx')






url='https://www.qdxq.sdu.edu.cn/xwzx1/xwjj/jcdt.htm'
resp=requests.get(url=url)
temp=resp.content.decode()
obj=re.compile(r"<li>.*?/../(?P<aaa>info/\w{4}/\w{5}.htm).*?<span>(?P<bbb>.*?)</span>.*?<span>(?P<ccc>.*?)</span>.*?\"title\">(?P<ddd>.*?)</p>.*?\"content\">(?P<eee>.*?)</p>.*?</div>",re.S)
pri=obj.findall(temp)
print(pri)
data={
"context":["website","date","month","title","main idea"],
"idea0":pri[0],
"idea1":pri[1],
"idea2":pri[2],
"idea3":pri[3],
"idea4":pri[4],
"idea5":pri[5],
"idea6":pri[6],
"idea7":pri[7],
"idea8":pri[8],
"idea9":pri[9]
    }
df=pd.DataFrame(data)
df1=df.T
for i in df:
    print(i)
df1.to_excel("基层动态.xlsx",index=False)
wb=openpyxl.load_workbook('基层动态.xlsx')
sheet=wb['Sheet1']
sheet['A12']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A12'].value)
sheet['A3']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A3'].value)
sheet['A4']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A4'].value)
sheet['A5']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A5'].value)
sheet['A6']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A6'].value)
sheet['A7']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A7'].value)
sheet['A8']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A8'].value)
sheet['A9']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A9'].value)
sheet['A10']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A10'].value)
sheet['A11']='https://www.qdxq.sdu.edu.cn/'+str(sheet['A11'].value)
wb.save('基层动态.xlsx')



url='https://www.qdxq.sdu.edu.cn/xwzx1/xwjj/xxtt.htm'
resp=requests.get(url=url)
temp=resp.content.decode()
obj=re.compile(r"<li>.*?<span>(?P<bbb>.*?)</span>.*?<span>(?P<ccc>.*?)</span>.*?\"title\">(?P<ddd>.*?)</p>.*?\"content\">(?P<eee>.*?)</p>.*?</div>",re.S)
pri=obj.findall(temp)
print(pri)
data={
"context":["date","month","title","main idea"],
"idea0":pri[0],
"idea1":pri[1],
"idea2":pri[2],
"idea3":pri[3],
"idea4":pri[4],
"idea5":pri[5],
"idea6":pri[6],
"idea7":pri[7],
"idea8":pri[8],
"idea9":pri[9]
    }
df=pd.DataFrame(data)
df1=df.T
for i in df:
    print(i)
df1.to_excel("学校头条.xlsx",index=False)

url='https://www.qdxq.sdu.edu.cn/xwzx1/tzyg/tzgg.htm'
resp=requests.get(url=url)
temp=resp.content.decode()
obj=re.compile(r"<li>.*?href=\"(?P<aaa>.*?)\".*?<span>(?P<bbb>.*?)</span>.*?<span>(?P<ccc>.*?)</span>.*?\"title\">(?P<ddd>.*?)</p>.*?\"content\">(?P<eee>.*?)</p>.*?</div>",re.S)
pri=obj.findall(temp)
print(pri)
data={
"context":["website","date","month","title","main idea"],
"idea0":pri[0],
"idea1":pri[1],
"idea2":pri[2],
"idea3":pri[3],
"idea4":pri[4],
"idea5":pri[5],
"idea6":pri[6],
"idea7":pri[7],
"idea8":pri[8],
"idea9":pri[9]
    }
df=pd.DataFrame(data)
df1=df.T
for i in df:
    print(i)
df1.to_excel("通知公告.xlsx",index=False)
wb=openpyxl.load_workbook('通知公告.xlsx')
sheet=wb['Sheet1']
i=3
while i<=12:
    if str(sheet[f'A{i}'].value[1])=='.':
        sheet[f'A{i}']='https://www.qdxq.sdu.edu.cn/'+str(sheet[f'A{i}'].value)
    i=i+1
wb.save('通知公告.xlsx')



url='https://www.qdxq.sdu.edu.cn/xwzx1/tzyg/xsyg.htm'
resp=requests.get(url=url)
temp=resp.content.decode()
obj=re.compile(r"<li>.*?href=\"(?P<aaa>.*?)\".*?<span>(?P<bbb>.*?)</span>.*?<span>(?P<ccc>.*?)</span>.*?\"title\">(?P<ddd>.*?)</p>.*?\"content\">(?P<eee>.*?)</p>.*?</div>",re.S)
pri=obj.findall(temp)
print(pri)
data={
"context":["website","date","month","title","main idea"],
"idea0":pri[0],
"idea1":pri[1],
"idea2":pri[2],
"idea3":pri[3],
"idea4":pri[4],
"idea5":pri[5],
"idea6":pri[6],
"idea7":pri[7],
"idea8":pri[8],
"idea9":pri[9]
    }
df=pd.DataFrame(data)
df1=df.T
for i in df:
    print(i)
df1.to_excel("学术预告.xlsx",index=False)
wb=openpyxl.load_workbook('学术预告.xlsx')
sheet=wb['Sheet1']
i=3
while i<=12:
    if str(sheet[f'A{i}'].value[1])=='.':
        sheet[f'A{i}']='https://www.qdxq.sdu.edu.cn/'+str(sheet[f'A{i}'].value)
    i=i+1
wb.save('学术预告.xlsx')



df1 = pd.read_excel('校区要闻.xlsx')
df2 = pd.read_excel('基层动态.xlsx')
df3 = pd.read_excel('学校头条.xlsx')
df4 = pd.read_excel('通知公告.xlsx')
df5 = pd.read_excel('学术预告.xlsx')
with pd.ExcelWriter('青岛校区新闻聚焦.xlsx') as writer:
    df1.to_excel(writer, sheet_name='校区要闻', index=False)
    df2.to_excel(writer, sheet_name='基层动态', index=False)
    df3.to_excel(writer, sheet_name='没有网址的学校头条', index=False)
    df4.to_excel(writer, sheet_name='通知公告', index=False)
    df5.to_excel(writer, sheet_name='学术预告', index=False)
os.remove('校区要闻.xlsx')
os.remove('学校头条.xlsx')
os.remove('基层动态.xlsx')
os.remove('通知公告.xlsx')
os.remove('学术预告.xlsx')